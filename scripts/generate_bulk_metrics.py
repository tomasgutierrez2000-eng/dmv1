#!/usr/bin/env python3
"""
Bulk YAML metric generator from Excel spreadsheet.
Reads Metrics-initial.xlsx and generates YAML metric definitions
for the GSIB calc engine pipeline.
"""
import json
import os
import re
import textwrap
from pathlib import Path

import openpyxl

# ── Configuration ──────────────────────────────────────────────
EXCEL_PATH = os.path.expanduser("~/Downloads/Metrics-initial.xlsx")
OUTPUT_DIR = Path(__file__).parent / "calc_engine" / "metrics"
DD_PATH = Path(__file__).parent.parent / "facility-summary-mvp" / "output" / "data-dictionary" / "data-dictionary.json"

# Domain mapping from Excel "Data Group" → YAML domain + sub_domain
DOMAIN_MAP = {
    "Reference": ("reference", "master-data"),
    "Pricing": ("pricing", "facility-pricing"),
    "Amendments": ("amendments", "amendment-events"),
    "Collateral": ("exposure", "collateral"),
    "Capital": ("capital", "adequacy"),
    "Captial": ("capital", "adequacy"),  # typo in Excel
    "Exposure, Profitability & Risk Metrics": ("exposure", "risk-metrics"),
    "Risk Ratings": ("risk", "credit-ratings"),
}

# Metric ID prefix from domain
DOMAIN_PREFIX = {
    "reference": "REF",
    "pricing": "PRC",
    "amendments": "AMD",
    "exposure": "EXP",
    "capital": "CAP",
    "risk": "RSK",
}

# Data type → unit_type + display_format
TYPE_MAP = {
    "Decimal": ("CURRENCY", "$,.0f"),
    "Integer": ("COUNT", ",.0f"),
    "String": ("ORDINAL", "s"),
    "Boolean": ("ORDINAL", "s"),
    "Date": ("ORDINAL", "s"),
}

# Special unit_type overrides by metric pattern
def infer_unit_type(name, field, data_type):
    nl = name.lower()
    fl = field.lower() if field else ""
    if "_pct" in fl or "(%)" in name or "ratio" in nl:
        return "PERCENTAGE", ".2f"
    if "_bps" in fl or "spread" in nl:
        return "BPS", ",.0f"
    if "count" in nl or "number" in nl:
        return "COUNT", ",.0f"
    if "_amt" in fl or "($)" in name or "amount" in nl or "income" in nl or "expense" in nl or "cost" in nl or "revenue" in nl:
        return "CURRENCY", "$,.0f"
    if "days" in nl or "tenor" in nl:
        return "DAYS", ",.0f"
    if "rate" in nl and "_pct" not in fl:
        return "RATE", ".4f"
    if "ratio" in nl:
        return "RATIO", ".2f"
    if "flag" in nl or "flag" in fl:
        return "ORDINAL", "s"
    if "bucket" in nl or "tier" in nl or "status" in nl or "type" in nl:
        return "ORDINAL", "s"
    if "date" in nl or "_date" in fl:
        return "ORDINAL", "s"
    if "id" in nl and fl.endswith("_id"):
        return "ORDINAL", "s"
    if "name" in fl or "code" in fl:
        return "ORDINAL", "s"
    return TYPE_MAP.get(data_type, ("CURRENCY", "$,.0f"))


def infer_direction(name, field):
    nl = name.lower()
    # Higher is better
    if any(x in nl for x in ["income", "revenue", "collateral value", "coverage", "net worth"]):
        return "HIGHER_BETTER"
    # Lower is better
    if any(x in nl for x in ["loss", "delinquen", "overdue", "cost", "expense", "risk weight",
                              "days past due", "deteriorat"]):
        return "LOWER_BETTER"
    return "NEUTRAL"


def infer_metric_class(sourcing_types):
    """Infer metric_class from level sourcing types."""
    types = set(t for t in sourcing_types if t)
    if types <= {"Raw"}:
        return "SOURCED"
    if "Calculation" in types or "Weighted Average" in types:
        return "CALCULATED"
    if "Aggregation" in types:
        return "HYBRID"
    return "CALCULATED"


# ── Source Table Patterns ──────────────────────────────────────
# Map field_name → best primary source table for SQL formulas
FIELD_SOURCE_MAP = {
    # L2 Exposure
    "number_of_loans": ("l2", "facility_exposure_snapshot", "fes", "INTEGER"),
    "drawn_amount": ("l2", "facility_exposure_snapshot", "fes", "NUMERIC"),
    "committed_amount": ("l2", "facility_exposure_snapshot", "fes", "NUMERIC"),
    "undrawn_amount": ("l2", "facility_exposure_snapshot", "fes", "NUMERIC"),
    "days_until_maturity": ("l2", "facility_exposure_snapshot", "fes", "INTEGER"),
    "facility_utilization_status": ("l2", "facility_exposure_snapshot", "fes", "VARCHAR"),
    "number_of_facilities": ("l2", "facility_exposure_snapshot", "fes", "INTEGER"),
    "rwa_amt": ("l2", "facility_exposure_snapshot", "fes", "NUMERIC"),
    "limit_status_code": ("l2", "facility_exposure_snapshot", "fes", "VARCHAR"),
    "coverage_ratio_pct": ("l2", "facility_exposure_snapshot", "fes", "NUMERIC"),
    "gross_exposure_usd": ("l2", "facility_exposure_snapshot", "fes", "NUMERIC"),
    "net_exposure_usd": ("l2", "facility_exposure_snapshot", "fes", "NUMERIC"),
    "outstanding_balance_amt": ("l2", "facility_exposure_snapshot", "fes", "NUMERIC"),
    "internal_risk_rating_bucket_code": ("l2", "facility_exposure_snapshot", "fes", "VARCHAR"),
    "total_collateral_mv_usd": ("l2", "facility_exposure_snapshot", "fes", "NUMERIC"),
    "fr2590_category_code": ("l2", "facility_exposure_snapshot", "fes", "VARCHAR"),

    # L2 Pricing
    "all_in_rate_pct": ("l2", "facility_pricing_snapshot", "fps", "NUMERIC"),
    "base_rate_pct": ("l2", "facility_pricing_snapshot", "fps", "NUMERIC"),
    "spread_bps": ("l2", "facility_pricing_snapshot", "fps", "NUMERIC"),
    "fee_rate_pct": ("l2", "facility_pricing_snapshot", "fps", "NUMERIC"),
    "pricing_tier": ("l2", "facility_pricing_snapshot", "fps", "VARCHAR"),
    "pricing_exception_flag": ("l2", "facility_pricing_snapshot", "fps", "CHAR"),
    "cost_of_funds_pct": ("l2", "facility_pricing_snapshot", "fps", "NUMERIC"),

    # L2 Financial
    "dscr_value": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),
    "ltv_pct": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),
    "noi_amt": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),
    "total_debt_service_amt": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),
    "revenue_amt": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),
    "operating_expense_amt": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),
    "ebitda_amt": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),
    "interest_expense_amt": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),
    "principal_payment_amt": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),
    "net_income_amt": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),
    "interest_rate_sensitivity_pct": ("l2", "facility_financial_snapshot", "ffs", "NUMERIC"),

    # L2 Profitability
    "interest_income_amt": ("l2", "facility_profitability_snapshot", "fprs", "NUMERIC"),
    "allocated_equity_amt": ("l2", "facility_profitability_snapshot", "fprs", "NUMERIC"),
    "avg_earning_assets_amt": ("l2", "facility_profitability_snapshot", "fprs", "NUMERIC"),

    # L2 Collateral
    "current_valuation_usd": ("l2", "collateral_snapshot", "cs", "NUMERIC"),
    "original_valuation_usd": ("l2", "collateral_snapshot", "cs", "NUMERIC"),
    "allocated_amount_usd": ("l2", "collateral_snapshot", "cs", "NUMERIC"),

    # L2 Risk
    "pd_pct": ("l2", "facility_risk_snapshot", "frs", "NUMERIC"),
    "lgd_pct": ("l2", "facility_risk_snapshot", "frs", "NUMERIC"),
    "ead_amt": ("l2", "facility_risk_snapshot", "frs", "NUMERIC"),
    "expected_loss_amt": ("l2", "facility_risk_snapshot", "frs", "NUMERIC"),
    "risk_weight_pct": ("l2", "facility_risk_snapshot", "frs", "NUMERIC"),

    # L2 Delinquency
    "delinquent_payment_flag": ("l2", "facility_delinquency_snapshot", "fds", "CHAR"),
    "days_past_due": ("l2", "facility_delinquency_snapshot", "fds", "INTEGER"),

    # L2 Ratings
    "risk_rating_status": ("l2", "counterparty_rating_observation", "cro", "VARCHAR"),
    "risk_rating_change_steps": ("l2", "counterparty_rating_observation", "cro", "INTEGER"),

    # L2 Position
    "lgd_estimate": ("l2", "position", "pos", "VARCHAR"),
    "funded_amount": ("l2", "position_detail", "pd", "NUMERIC"),
    "unfunded_amount": ("l2", "position_detail", "pd", "NUMERIC"),
    "contractual_amt": ("l2", "cash_flow", "cf", "NUMERIC"),
    "exposure_at_default": ("l2", "credit_event_facility_link", "cefl", "NUMERIC"),

    # L2 Deal pipeline
    "expected_tenor_months": ("l2", "deal_pipeline_fact", "dpf", "VARCHAR"),

    # L2 Exception
    "exception_status": ("l2", "exception_event", "ee", "VARCHAR"),

    # L2 Amendment
    "amendment_event_id": ("l2", "amendment_event", "ae", "BIGINT"),

    # L1 Facility Master
    "facility_id": ("l1", "facility_master", "fm", "BIGINT"),
    "facility_name": ("l1", "facility_master", "fm", "VARCHAR"),
    "facility_code": ("l1", "facility_master", "fm", "VARCHAR"),
    "committed_facility_amt": ("l1", "facility_master", "fm", "NUMERIC"),
    "maturity_date": ("l1", "facility_master", "fm", "DATE"),
    "active_flag": ("l1", "facility_master", "fm", "CHAR"),
    "lob_segment_id": ("l1", "facility_master", "fm", "BIGINT"),
    "product_node_id": ("l1", "facility_master", "fm", "BIGINT"),
    "origination_date": ("l1", "facility_master", "fm", "DATE"),
    "facility_type_code": ("l1", "facility_master", "fm", "VARCHAR"),
    "interest_rate_spread_bps": ("l1", "facility_master", "fm", "NUMERIC"),

    # L1 Counterparty
    "counterparty_id": ("l1", "counterparty", "cp", "BIGINT"),
    "legal_name": ("l1", "counterparty", "cp", "VARCHAR"),
    "internal_risk_rating": ("l1", "counterparty", "cp", "VARCHAR"),
    "external_risk_rating": ("l2", "position", "pos", "VARCHAR"),
    "region_code": ("l1", "counterparty", "cp", "VARCHAR"),

    # L1 Credit Agreement
    "credit_agreement_id": ("l1", "credit_agreement_master", "cam", "BIGINT"),

    # L1 Dimensions
    "country_name": ("l1", "country_dim", "cd", "VARCHAR"),
    "industry_name": ("l1", "industry_dim", "ind", "VARCHAR"),
    "amendment_status_name": ("l1", "amendment_status_dim", "asd", "VARCHAR"),
    "amendment_type_code": ("l1", "amendment_type_dim", "atd", "VARCHAR"),
    "as_of_date": ("l2", "facility_exposure_snapshot", "fes", "DATE"),

    # L1 LoB
    "parent": ("l1", "enterprise_business_taxonomy", "ebt", "BIGINT"),

    # L1/L2 Participation
    "allocation_pct": ("l1", "facility_counterparty_participation", "fcp", "NUMERIC"),

    # L2 Limit
    "utilized_amount_usd": ("l2", "limit_utilization_event", "lue", "NUMERIC"),

    # L2 Counterparty Financial
    "total_assets_amt": ("l2", "counterparty_financial_snapshot", "cfs", "NUMERIC"),

    # L1 Risk mitigant
    "risk_mitigant_subtype_code": ("l1", "risk_mitigant_type_dim", "rmtd", "VARCHAR"),
    "risk_mitigant_id": ("l1", "risk_mitigant_master", "rmm", "BIGINT"),

    # L1 Maturity bucket
    "maturity_bucket_id": ("l1", "maturity_bucket_dim", "mbd", "BIGINT"),

    # L3 Derived
    "criticized_portfolio_count": ("l3", "lob_credit_quality_summary", "lcqs", "INTEGER"),
    "delinquency_rate_pct": ("l3", "lob_delinquency_summary", "lds", "NUMERIC"),
    "total_overdue_amt": ("l3", "lob_delinquency_summary", "lds", "NUMERIC"),
    "exception_rate_pct": ("l3", "lob_risk_ratio_summary", "lrrs", "NUMERIC"),
    "expected_loss_usd": ("l3", "stress_test_result", "str_res", "NUMERIC"),
    "limit_amt": ("l3", "lob_exposure_summary", "les", "NUMERIC"),
    "has_cross_entity_flag": ("l3", "counterparty_exposure_summary", "ces", "CHAR"),
    "is_syndicated_flag": ("l3", "facility_detail_snapshot", "fds", "CHAR"),
    "is_deteriorated": ("l3", "facility_detail_snapshot", "fds", "CHAR"),  # to be added
    "utilization_pct": ("l3", "lob_exposure_summary", "les", "NUMERIC"),
    "nim_pct": ("l3", "lob_profitability_summary", "lps", "NUMERIC"),
    "roa_pct": ("l3", "lob_profitability_summary", "lps", "NUMERIC"),
    "roe_pct": ("l3", "lob_profitability_summary", "lps", "NUMERIC"),
    "rwa_density_pct": ("l3", "lob_exposure_summary", "les", "NUMERIC"),
    "fccr_value": ("l3", "lob_risk_ratio_summary", "lrrs", "NUMERIC"),
    "capital_adequacy_ratio_pct": ("l3", "lob_risk_ratio_summary", "lrrs", "NUMERIC"),
    "origination_date_bucket": ("l3", "facility_detail_snapshot", "fds", "VARCHAR"),
    "effective_date_bucket": ("l3", "facility_detail_snapshot", "fds", "VARCHAR"),
    "facility_active_flag": ("l3", "facility_detail_snapshot", "fds", "CHAR"),

    # Missing - will be added
    "total_cross_entity_exposure_usd": ("l3", "counterparty_exposure_summary", "ces", "NUMERIC"),
    "expected_loss_rate_pct": ("l2", "facility_risk_snapshot", "frs", "NUMERIC"),
    "tangible_net_worth_usd": ("l2", "counterparty_financial_snapshot", "cfs", "NUMERIC"),
    "return_on_rwa_pct": ("l3", "lob_profitability_summary", "lps", "NUMERIC"),
    "dpd_bucket_code": ("l2", "facility_delinquency_snapshot", "fds", "VARCHAR"),
}

# ── Existing metrics to skip (already have YAMLs) ─────────────
EXISTING_METRIC_FIELDS = {
    # EXP-001 = Current Collateral Market Value = current_valuation_usd (ref 11)
    # PROF-108 = Interest Expense = interest_expense_amt (ref 86)
}
# We'll generate these too but with their existing IDs

# ── SQL Pattern Builders ───────────────────────────────────────

def make_facility_sql_raw(schema, table, alias, field, join_fm=True):
    """Raw field lookup at facility level."""
    if schema == "l1" and table == "facility_master":
        return f"""\
SELECT
  fm.facility_id AS dimension_key,
  fm.{field} AS metric_value
FROM l1.facility_master fm
WHERE fm.active_flag = 'Y'
  AND fm.is_current_flag = 'Y'"""
    elif schema == "l1" and table == "counterparty":
        return f"""\
SELECT
  fm.facility_id AS dimension_key,
  cp.{field} AS metric_value
FROM l1.facility_master fm
INNER JOIN l1.counterparty cp
  ON cp.counterparty_id = fm.counterparty_id
  AND cp.is_current_flag = 'Y'
WHERE fm.active_flag = 'Y'
  AND fm.is_current_flag = 'Y'"""
    elif schema == "l1":
        return f"""\
SELECT
  fm.facility_id AS dimension_key,
  {alias}.{field} AS metric_value
FROM l1.facility_master fm
INNER JOIN {schema}.{table} {alias}
  ON {alias}.facility_id = fm.facility_id
WHERE fm.active_flag = 'Y'"""
    elif schema == "l2":
        return f"""\
SELECT
  {alias}.facility_id AS dimension_key,
  {alias}.{field} AS metric_value
FROM {schema}.{table} {alias}
WHERE {alias}.as_of_date = :as_of_date"""
    return ""


def make_facility_sql_sum(schema, table, alias, field, bank_share=False):
    """SUM aggregation at facility level."""
    bs = "\n    * COALESCE(fla.bank_share_pct, 100.0) / 100.0" if bank_share else ""
    fla_join = """\
LEFT JOIN l1.facility_lender_allocation fla
  ON fla.facility_id = {alias}.facility_id
  AND fla.is_current_flag = 'Y'""".format(alias=alias) if bank_share else ""

    return f"""\
SELECT
  {alias}.facility_id AS dimension_key,
  SUM({alias}.{field}{bs}) AS metric_value
FROM {schema}.{table} {alias}
INNER JOIN l1.facility_master fm
  ON fm.facility_id = {alias}.facility_id
  AND fm.active_flag = 'Y'
{fla_join}
WHERE {alias}.as_of_date = :as_of_date
GROUP BY {alias}.facility_id"""


def make_facility_sql_count(schema, table, alias, field):
    """COUNT at facility level."""
    return f"""\
SELECT
  {alias}.facility_id AS dimension_key,
  COUNT(DISTINCT {alias}.{field}) AS metric_value
FROM {schema}.{table} {alias}
WHERE {alias}.as_of_date = :as_of_date
GROUP BY {alias}.facility_id"""


def make_facility_sql_wavg(schema, table, alias, field, weight_field="committed_amount"):
    """Weighted average at facility level."""
    return f"""\
SELECT
  {alias}.facility_id AS dimension_key,
  SUM({alias}.{field} * {alias}.{weight_field})
    / NULLIF(SUM({alias}.{weight_field}), 0) AS metric_value
FROM {schema}.{table} {alias}
WHERE {alias}.as_of_date = :as_of_date
GROUP BY {alias}.facility_id"""


def make_facility_sql_calc(formula_expr, schema, table, alias, extra_joins=""):
    """Custom calculation at facility level."""
    return f"""\
SELECT
  {alias}.facility_id AS dimension_key,
  {formula_expr} AS metric_value
FROM {schema}.{table} {alias}
INNER JOIN l1.facility_master fm
  ON fm.facility_id = {alias}.facility_id
  AND fm.active_flag = 'Y'
{extra_joins}
WHERE {alias}.as_of_date = :as_of_date
GROUP BY {alias}.facility_id"""


def make_counterparty_sql(fac_sql_inner, agg_func="SUM"):
    """Counterparty level: aggregate facility-level values."""
    return f"""\
SELECT
  fm.counterparty_id AS dimension_key,
  {agg_func}(fac.metric_value) AS metric_value
FROM (
{textwrap.indent(fac_sql_inner, '  ')}
) fac
INNER JOIN l1.facility_master fm
  ON fm.facility_id = fac.dimension_key
  AND fm.active_flag = 'Y'
GROUP BY fm.counterparty_id"""


def make_counterparty_sql_raw(schema, table, alias, field):
    """Raw counterparty lookup."""
    if schema == "l1" and table == "counterparty":
        return f"""\
SELECT
  cp.counterparty_id AS dimension_key,
  cp.{field} AS metric_value
FROM l1.counterparty cp
WHERE cp.is_current_flag = 'Y'"""
    return f"""\
SELECT
  fm.counterparty_id AS dimension_key,
  {alias}.{field} AS metric_value
FROM {schema}.{table} {alias}
INNER JOIN l1.facility_master fm
  ON fm.facility_id = {alias}.facility_id
  AND fm.active_flag = 'Y'
WHERE {alias}.as_of_date = :as_of_date"""


def make_desk_sql(fac_sql_inner, agg_func="SUM"):
    """Desk (L3 LoB) level: aggregate through enterprise_business_taxonomy."""
    return f"""\
SELECT
  ebt.managed_segment_id AS dimension_key,
  {agg_func}(fac.metric_value) AS metric_value
FROM (
{textwrap.indent(fac_sql_inner, '  ')}
) fac
INNER JOIN l1.facility_master fm
  ON fm.facility_id = fac.dimension_key
  AND fm.active_flag = 'Y'
LEFT JOIN l1.enterprise_business_taxonomy ebt
  ON ebt.managed_segment_id = fm.lob_segment_id
GROUP BY ebt.managed_segment_id"""


def make_portfolio_sql(fac_sql_inner, agg_func="SUM"):
    """Portfolio (L2 LoB) level."""
    return f"""\
SELECT
  ebt_l2.managed_segment_id AS dimension_key,
  {agg_func}(fac.metric_value) AS metric_value
FROM (
{textwrap.indent(fac_sql_inner, '  ')}
) fac
INNER JOIN l1.facility_master fm
  ON fm.facility_id = fac.dimension_key
  AND fm.active_flag = 'Y'
LEFT JOIN l1.enterprise_business_taxonomy ebt_l3
  ON ebt_l3.managed_segment_id = fm.lob_segment_id
LEFT JOIN l1.enterprise_business_taxonomy ebt_l2
  ON ebt_l2.managed_segment_id = ebt_l3.parent_segment_id
GROUP BY ebt_l2.managed_segment_id"""


def make_bizseg_sql(fac_sql_inner, agg_func="SUM"):
    """Business segment (L1 LoB) level."""
    return f"""\
SELECT
  ebt_l1.managed_segment_id AS dimension_key,
  {agg_func}(fac.metric_value) AS metric_value
FROM (
{textwrap.indent(fac_sql_inner, '  ')}
) fac
INNER JOIN l1.facility_master fm
  ON fm.facility_id = fac.dimension_key
  AND fm.active_flag = 'Y'
LEFT JOIN l1.enterprise_business_taxonomy ebt_l3
  ON ebt_l3.managed_segment_id = fm.lob_segment_id
LEFT JOIN l1.enterprise_business_taxonomy ebt_l2
  ON ebt_l2.managed_segment_id = ebt_l3.parent_segment_id
LEFT JOIN l1.enterprise_business_taxonomy ebt_l1
  ON ebt_l1.managed_segment_id = ebt_l2.parent_segment_id
GROUP BY ebt_l1.managed_segment_id"""


# ── Aggregation type inference ─────────────────────────────────
def infer_agg_type(sourcing_type):
    mapping = {
        "Raw": "RAW",
        "Aggregation": "SUM",
        "Calculation": "CUSTOM",
        "Average": "WEIGHTED_AVG",
        "Weighted Average": "WEIGHTED_AVG",
    }
    return mapping.get(sourcing_type, "SUM")


def map_agg_func(agg_type):
    """Map aggregation type to SQL function."""
    if agg_type in ("WEIGHTED_AVG",):
        return "AVG"
    if agg_type == "COUNT" or agg_type == "COUNT_DISTINCT":
        return "SUM"  # SUM of facility-level counts
    return "SUM"


# ── Source table builder ───────────────────────────────────────
def build_source_tables(field, schema, table, alias):
    """Build YAML source_tables block."""
    tables = []

    # Base table
    base = {
        "schema": schema,
        "table": table,
        "alias": alias,
        "join_type": "BASE",
        "fields": [],
    }

    # Add standard fields
    if schema == "l2" or (schema == "l1" and table != "facility_master"):
        base["fields"].append({"name": "facility_id", "role": "JOIN_KEY"})
    if schema == "l2":
        base["fields"].append({"name": "as_of_date", "role": "FILTER"})
    if schema == "l1" and table == "facility_master":
        base["fields"].append({"name": "facility_id", "role": "JOIN_KEY"})
        base["fields"].append({"name": "active_flag", "role": "FILTER"})
    if field != "facility_id" and field != "as_of_date":
        base["fields"].append({"name": field, "role": "MEASURE"})

    tables.append(base)

    # Add facility_master join if base is L2
    if schema == "l2":
        fm = {
            "schema": "l1",
            "table": "facility_master",
            "alias": "fm",
            "join_type": "INNER",
            "join_on": f"fm.facility_id = {alias}.facility_id",
            "fields": [
                {"name": "facility_id", "role": "JOIN_KEY"},
                {"name": "counterparty_id", "role": "DIMENSION"},
                {"name": "lob_segment_id", "role": "DIMENSION"},
                {"name": "active_flag", "role": "FILTER"},
            ],
        }
        tables.append(fm)

    # Add EBT for rollup
    ebt = {
        "schema": "l1",
        "table": "enterprise_business_taxonomy",
        "alias": "ebt",
        "join_type": "LEFT",
        "join_on": "ebt.managed_segment_id = fm.lob_segment_id",
        "fields": [
            {"name": "managed_segment_id", "role": "DIMENSION"},
            {"name": "parent_segment_id", "role": "DIMENSION"},
        ],
    }
    tables.append(ebt)

    return tables


# ── YAML Writer ────────────────────────────────────────────────
def write_yaml(metric_def, output_path):
    """Write a metric definition to YAML file."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    m = metric_def

    lines = []
    lines.append(f'# {"═" * 63}')
    lines.append(f'# GSIB Metric Definition')
    lines.append(f'# {m["name"]}')
    lines.append(f'# {"═" * 63}')
    lines.append(f'# ── IDENTIFICATION {"─" * 44}')
    lines.append(f'metric_id: "{m["metric_id"]}"')
    lines.append(f'name: "{m["name"]}"')
    lines.append(f'version: "1.0.0"')
    lines.append(f'owner: "{m["owner"]}"')
    lines.append(f'status: DRAFT')
    lines.append(f'effective_date: "2026-01-01"')
    lines.append(f'supersedes: null')

    lines.append(f'# ── CLASSIFICATION {"─" * 44}')
    lines.append(f'domain: "{m["domain"]}"')
    lines.append(f'sub_domain: "{m["sub_domain"]}"')
    lines.append(f'metric_class: {m["metric_class"]}')
    lines.append(f'direction: {m["direction"]}')
    lines.append(f'unit_type: {m["unit_type"]}')
    lines.append(f'display_format: "{m["display_format"]}"')
    # Multi-line description
    lines.append(f'description: >')
    for dl in textwrap.wrap(m["description"], 78):
        lines.append(f'  {dl}')

    lines.append(f'# ── REGULATORY REFERENCES {"─" * 37}')
    lines.append(f'regulatory_references:')
    for ref in m.get("regulatory_references", [{"framework": "Basel III", "section": "", "description": ""}]):
        lines.append(f'  - framework: "{ref["framework"]}"')
        if ref.get("section"):
            lines.append(f'    section: "{ref["section"]}"')
        if ref.get("schedule"):
            lines.append(f'    schedule: "{ref["schedule"]}"')
        lines.append(f'    description: "{ref.get("description", "")}"')

    lines.append(f'# ── SOURCE TABLES {"─" * 45}')
    lines.append(f'source_tables:')
    for st in m["source_tables"]:
        lines.append(f'  - schema: {st["schema"]}')
        lines.append(f'    table: {st["table"]}')
        lines.append(f'    alias: {st["alias"]}')
        lines.append(f'    join_type: {st["join_type"]}')
        if st.get("join_on"):
            lines.append(f'    join_on: "{st["join_on"]}"')
        lines.append(f'    fields:')
        for fld in st["fields"]:
            lines.append(f'      - name: {fld["name"]}')
            lines.append(f'        role: {fld["role"]}')
            if fld.get("description"):
                lines.append(f'        description: "{fld["description"]}"')

    lines.append(f'# ── LEVEL FORMULAS {"─" * 44}')
    lines.append(f'levels:')
    for level_name in ["facility", "counterparty", "desk", "portfolio", "business_segment"]:
        level = m["levels"][level_name]
        lines.append(f'  {level_name}:')
        lines.append(f'    aggregation_type: {level["aggregation_type"]}')
        lines.append(f'    formula_text: >')
        for fl in textwrap.wrap(level["formula_text"], 76):
            lines.append(f'      {fl}')
        lines.append(f'    formula_sql: |')
        for sl in level["formula_sql"].rstrip().split('\n'):
            lines.append(f'      {sl}')

    lines.append(f'# ── DEPENDENCIES {"─" * 46}')
    lines.append(f'depends_on: {json.dumps(m.get("depends_on", []))}')

    lines.append(f'# ── OUTPUT {"─" * 52}')
    lines.append(f'output:')
    lines.append(f'  table: metric_result')

    lines.append(f'# ── VALIDATION RULES {"─" * 42}')
    lines.append(f'validations:')
    for v in m.get("validations", []):
        lines.append(f'  - rule_id: "{v["rule_id"]}"')
        lines.append(f'    type: {v["type"]}')
        lines.append(f'    description: "{v["description"]}"')
        lines.append(f'    severity: {v["severity"]}')

    lines.append(f'# ── CATALOGUE {"─" * 49}')
    lines.append(f'catalogue:')
    lines.append(f'  item_id: "{m["catalogue_item_id"]}"')
    lines.append(f'  abbreviation: "{m["abbreviation"]}"')
    lines.append(f'  insight: >')
    for il in textwrap.wrap(m.get("insight", ""), 76):
        lines.append(f'    {il}')
    lines.append(f'  rollup_strategy: "{m.get("rollup_strategy", "direct-sum")}"')
    lines.append(f'  primary_value_field: "{m["field_name"]}"')

    lines.append(f'# ── METADATA {"─" * 50}')
    lines.append(f'tags: {json.dumps(m.get("tags", []))}')
    lines.append(f'dashboard_pages: {json.dumps(m.get("dashboard_pages", []))}')
    lines.append(f'legacy_metric_ids: []')

    with open(output_path, 'w') as f:
        f.write('\n'.join(lines) + '\n')

    return output_path


# ── Main Generator ─────────────────────────────────────────────
def parse_excel():
    """Parse Excel metrics spreadsheet."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Sheet1']

    metrics = []
    for r in range(3, ws.max_row + 1):
        def v(c):
            val = ws.cell(r, c).value
            return str(val).strip() if val is not None else None

        metric = {
            'ref': v(1),
            'status': v(2),
            'owner': v(6),
            'data_group': v(7),
            'name': v(8),
            'field_name': v(9),
            'layer': v(10),
            'definition': v(11),
            'example_kpi': v(12),
            'num_instances': v(13),
            'insight': v(15),
            'data_type': v(16),
            'levels': {
                'facility': {'in_record': v(19), 'sourcing_type': v(20), 'level_logic': v(21)},
                'counterparty': {'in_record': v(23), 'sourcing_type': v(24), 'level_logic': v(25)},
                'desk': {'in_record': v(27), 'sourcing_type': v(28), 'level_logic': v(29)},
                'portfolio': {'in_record': v(31), 'sourcing_type': v(32), 'level_logic': v(33)},
                'business_segment': {'in_record': v(35), 'sourcing_type': v(36), 'level_logic': v(37)},
            },
        }
        metrics.append(metric)

    return metrics


def generate_metric_yaml(m, metric_id_counters):
    """Generate a complete YAML metric definition from an Excel row."""
    ref = m['ref']
    name = m['name']
    if not name:
        return None

    # Clean up name (remove quotes and brackets)
    clean_name = name.replace('"', '').strip()
    # Remove content in square brackets from display but keep for context
    display_name = re.sub(r'\s*\[.*?\]', '', clean_name).strip()
    if display_name.startswith('"') or display_name.endswith('"'):
        display_name = display_name.strip('"')

    # Field name
    field = m['field_name']
    if not field:
        # Auto-assign for missing fields
        if "days payment" in name.lower() and "overdue" in name.lower():
            field = "dpd_bucket_code"
        else:
            field = re.sub(r'[^a-z0-9]+', '_', name.lower()).strip('_')

    field_clean = field.lower().strip()
    if '(' in field_clean:
        field_clean = field_clean.split('(')[0].strip()

    # Fix typos
    if field_clean == "interest_risk_rating_bucket_code":
        field_clean = "internal_risk_rating_bucket_code"

    # Domain
    data_group = m['data_group'] or "Reference"
    domain, sub_domain = DOMAIN_MAP.get(data_group, ("reference", "general"))

    # Metric ID
    prefix = DOMAIN_PREFIX.get(domain, "MET")
    if prefix not in metric_id_counters:
        metric_id_counters[prefix] = 0
    metric_id_counters[prefix] += 1
    metric_id = f"{prefix}-{metric_id_counters[prefix]:03d}"

    # Catalogue item ID
    catalogue_item_id = f"MET-{int(ref):03d}" if ref and ref.isdigit() else f"MET-{metric_id_counters[prefix]:03d}"

    # Data type inference
    data_type = m['data_type'] or "Decimal"
    unit_type, display_format = infer_unit_type(name, field_clean, data_type)
    direction = infer_direction(name, field_clean)
    sourcing_types = [m['levels'][l]['sourcing_type'] for l in m['levels']]
    metric_class = infer_metric_class(sourcing_types)

    # Source table lookup
    src_info = FIELD_SOURCE_MAP.get(field_clean)
    if not src_info:
        print(f"  WARNING: No source mapping for {field_clean} (ref {ref})")
        src_info = ("l2", "facility_exposure_snapshot", "fes", "VARCHAR")

    src_schema, src_table, src_alias, src_type = src_info

    # Build abbreviation
    words = re.sub(r'[^A-Z\s]', '', display_name.upper()).split()
    abbreviation = ''.join(w[0] for w in words[:4]) if words else field_clean[:8].upper()

    # Owner mapping
    owner_map = {
        "Clarke": "credit-risk",
        "Devi": "market-risk",
        None: "credit-risk",
    }
    owner = owner_map.get(m['owner'], "credit-risk")

    # ── Build level formulas ───────────────────────────────
    levels = {}

    # Determine facility-level SQL pattern
    fac_sourcing = m['levels']['facility']['sourcing_type'] or 'Raw'
    cp_sourcing = m['levels']['counterparty']['sourcing_type'] or fac_sourcing

    # Check if this is a monetary amount that should use bank_share
    uses_bank_share = unit_type == "CURRENCY" and "_amt" in field_clean or field_clean.endswith("_usd")

    if fac_sourcing == "Raw" and src_schema == "l1":
        fac_sql = make_facility_sql_raw(src_schema, src_table, src_alias, field_clean)
        fac_text = f"Direct lookup of {field_clean} from {src_table}."
        fac_agg = "RAW"
    elif fac_sourcing == "Raw":
        fac_sql = make_facility_sql_raw(src_schema, src_table, src_alias, field_clean)
        fac_text = f"Read {field_clean} from {src_table} for each facility as-of reporting date."
        fac_agg = "RAW"
    elif fac_sourcing == "Aggregation" and unit_type == "COUNT":
        fac_sql = make_facility_sql_count(src_schema, src_table, src_alias, field_clean)
        fac_text = f"COUNT DISTINCT {field_clean} per facility from {src_table}."
        fac_agg = "COUNT_DISTINCT"
    elif fac_sourcing == "Aggregation":
        fac_sql = make_facility_sql_sum(src_schema, src_table, src_alias, field_clean, bank_share=uses_bank_share)
        fac_text = f"SUM({field_clean}) per facility from {src_table}."
        if uses_bank_share:
            fac_text += " Weighted by bank share percentage."
        fac_agg = "SUM"
    elif fac_sourcing == "Weighted Average":
        weight = "committed_amount" if "committed_amount" in FIELD_SOURCE_MAP else "drawn_amount"
        fac_sql = make_facility_sql_wavg(src_schema, src_table, src_alias, field_clean, weight)
        fac_text = f"Exposure-weighted average of {field_clean} per facility."
        fac_agg = "WEIGHTED_AVG"
    elif fac_sourcing == "Calculation":
        # Use level_logic from Excel if available, otherwise generate
        logic = m['levels']['facility']['level_logic'] or ""
        fac_sql = make_facility_sql_raw(src_schema, src_table, src_alias, field_clean)
        fac_text = logic if logic else f"Calculated {field_clean} per facility."
        fac_agg = "CUSTOM"
    elif fac_sourcing == "Average":
        fac_sql = make_facility_sql_raw(src_schema, src_table, src_alias, field_clean)
        fac_text = f"Average of {field_clean} per facility."
        fac_agg = "WEIGHTED_AVG"
    else:
        fac_sql = make_facility_sql_raw(src_schema, src_table, src_alias, field_clean)
        fac_text = f"Direct {field_clean} from {src_table}."
        fac_agg = "RAW"

    levels["facility"] = {
        "aggregation_type": fac_agg,
        "formula_text": fac_text,
        "formula_sql": fac_sql,
    }

    # Counterparty level
    cp_in = m['levels']['counterparty']['in_record'] == 'Y'
    cp_type = m['levels']['counterparty']['sourcing_type'] or fac_sourcing

    if not cp_in:
        # Still generate SQL for completeness
        cp_agg_func = map_agg_func(fac_agg)
        cp_sql = make_counterparty_sql(fac_sql, cp_agg_func)
        cp_text = f"{cp_agg_func}(facility-level {field_clean}) per counterparty."
        cp_agg = "SUM"
    elif cp_type == "Raw" and src_schema == "l1" and src_table == "counterparty":
        cp_sql = make_counterparty_sql_raw(src_schema, src_table, src_alias, field_clean)
        cp_text = f"Direct lookup of {field_clean} from counterparty."
        cp_agg = "RAW"
    elif cp_type in ("Aggregation", "Calculation"):
        cp_agg_func = "SUM"
        cp_sql = make_counterparty_sql(fac_sql, cp_agg_func)
        cp_text = f"SUM of facility-level {field_clean} per counterparty."
        cp_agg = "SUM"
    elif cp_type in ("Average", "Weighted Average"):
        cp_sql = make_counterparty_sql(fac_sql, "AVG")
        cp_text = f"Exposure-weighted average of {field_clean} across counterparty facilities."
        cp_agg = "WEIGHTED_AVG"
    elif cp_type == "Raw":
        cp_sql = make_counterparty_sql_raw(src_schema, src_table, src_alias, field_clean)
        cp_text = f"Read {field_clean} per counterparty."
        cp_agg = "RAW"
    else:
        cp_sql = make_counterparty_sql(fac_sql, "SUM")
        cp_text = f"SUM of facility-level {field_clean} per counterparty."
        cp_agg = "SUM"

    levels["counterparty"] = {
        "aggregation_type": cp_agg,
        "formula_text": cp_text,
        "formula_sql": cp_sql,
    }

    # Desk, Portfolio, Business Segment — use standard LOB rollup
    for level_name, make_fn, lob_label in [
        ("desk", make_desk_sql, "L3 desk"),
        ("portfolio", make_portfolio_sql, "L2 portfolio"),
        ("business_segment", make_bizseg_sql, "L1 business segment"),
    ]:
        lvl = m['levels'][level_name]
        in_record = lvl['in_record'] == 'Y'
        st = lvl['sourcing_type'] or fac_sourcing

        if st in ("Aggregation", "Calculation"):
            agg_func = "SUM"
            agg_type = "SUM"
        elif st in ("Average", "Weighted Average"):
            agg_func = "AVG"
            agg_type = "WEIGHTED_AVG"
        else:
            agg_func = "SUM"
            agg_type = "SUM"

        lob_sql = make_fn(fac_sql, agg_func)
        lob_text = f"{agg_func} of facility-level {field_clean} per {lob_label} segment."

        levels[level_name] = {
            "aggregation_type": agg_type,
            "formula_text": lob_text,
            "formula_sql": lob_sql,
        }

    # Build source tables
    source_tables = build_source_tables(field_clean, src_schema, src_table, src_alias)

    # Build validation rules
    validations = [
        {"rule_id": f"{metric_id}-V01", "type": "NOT_NULL", "description": f"No null {field_clean} values", "severity": "ERROR"},
    ]
    if unit_type in ("CURRENCY", "COUNT"):
        validations.append(
            {"rule_id": f"{metric_id}-V02", "type": "NON_NEGATIVE", "description": f"{field_clean} must be >= 0", "severity": "ERROR"}
        )

    # Tags
    tags = [domain]
    if sub_domain != domain:
        tags.append(sub_domain)
    if "gsib" in (m['definition'] or "").lower():
        tags.append("gsib")

    # Build the full definition
    definition = {
        "metric_id": metric_id,
        "name": display_name,
        "owner": owner,
        "domain": domain,
        "sub_domain": sub_domain,
        "metric_class": metric_class,
        "direction": direction,
        "unit_type": unit_type,
        "display_format": display_format,
        "description": m['definition'] or f"{display_name} metric.",
        "regulatory_references": [{"framework": "Basel III", "section": "", "description": ""}],
        "source_tables": source_tables,
        "levels": levels,
        "depends_on": [],
        "validations": validations,
        "catalogue_item_id": catalogue_item_id,
        "abbreviation": abbreviation,
        "insight": m['insight'] or "",
        "rollup_strategy": "direct-sum" if unit_type in ("CURRENCY", "COUNT") else "weighted-avg" if unit_type in ("PERCENTAGE", "BPS", "RATIO", "RATE") else "direct-sum",
        "field_name": field_clean,
        "tags": tags,
        "dashboard_pages": [],
        "ref": ref,
    }

    return definition


def main():
    print("Parsing Excel...")
    metrics = parse_excel()
    print(f"Found {len(metrics)} metrics")

    metric_id_counters = {}
    generated = []
    corrections = []

    for m in metrics:
        if not m['name']:
            continue

        # Auto-corrections
        if not m['data_type']:
            name_lower = m['name'].lower()
            if 'ratio' in name_lower or '(%)' in name_lower or 'rate' in name_lower:
                m['data_type'] = 'Decimal'
                corrections.append(f"Ref {m['ref']}: Inferred data_type=Decimal for {m['name']}")
            elif 'amount' in name_lower or '($)' in name_lower:
                m['data_type'] = 'Decimal'
                corrections.append(f"Ref {m['ref']}: Inferred data_type=Decimal for {m['name']}")
            elif 'flag' in name_lower:
                m['data_type'] = 'Boolean'
                corrections.append(f"Ref {m['ref']}: Inferred data_type=Boolean for {m['name']}")
            else:
                m['data_type'] = 'String'
                corrections.append(f"Ref {m['ref']}: Inferred data_type=String for {m['name']}")

        if m['data_group'] == 'Captial':
            m['data_group'] = 'Capital'
            corrections.append(f"Ref {m['ref']}: Fixed typo Captial→Capital")

        if m['field_name'] and 'interest_risk_rating' in m['field_name']:
            old = m['field_name']
            m['field_name'] = m['field_name'].replace('interest_risk_rating', 'internal_risk_rating')
            corrections.append(f"Ref {m['ref']}: Fixed field name {old}→{m['field_name']}")

        # Layer corrections
        layer = m.get('layer', '') or ''
        field = (m['field_name'] or '').lower()
        if 'Reference' in layer and any(x in field for x in ['_pct', '_ratio', 'dscr', 'ltv']):
            corrections.append(f"Ref {m['ref']}: {m['name']} marked as Reference but field {field} looks Derived — keeping as-is for YAML")

        result = generate_metric_yaml(m, metric_id_counters)
        if result:
            generated.append(result)

    # Write YAML files
    print(f"\nGenerating {len(generated)} YAML files...")
    domain_dirs = {
        "reference": "reference",
        "pricing": "pricing",
        "amendments": "amendments",
        "exposure": "exposure",
        "capital": "capital",
        "risk": "risk",
    }

    written_files = []
    for defn in generated:
        domain = defn["domain"]
        dir_name = domain_dirs.get(domain, domain)
        out_dir = OUTPUT_DIR / dir_name
        filename = f"{defn['metric_id']}.yaml"
        out_path = out_dir / filename
        write_yaml(defn, str(out_path))
        written_files.append(str(out_path))
        print(f"  ✓ {out_path.relative_to(OUTPUT_DIR)}")

    # Print corrections summary
    print(f"\n{'=' * 60}")
    print(f"GENERATION COMPLETE")
    print(f"{'=' * 60}")
    print(f"Total metrics: {len(generated)}")
    print(f"Files written: {len(written_files)}")
    print(f"Auto-corrections: {len(corrections)}")
    for c in corrections:
        print(f"  • {c}")

    # Print metric ID mapping
    print(f"\n{'=' * 60}")
    print(f"METRIC ID MAPPING")
    print(f"{'=' * 60}")
    for defn in generated:
        print(f"  Ref {defn['ref']:>3s}: {defn['metric_id']:>8s} → {defn['catalogue_item_id']:>7s} | {defn['name']}")

    return generated, corrections


if __name__ == "__main__":
    main()
