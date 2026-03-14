#!/usr/bin/env python3
"""Load L1 reference data from Excel into PostgreSQL.
Replaces data in 49 L1 tables from colored Excel tabs.

Usage:
  python3 scripts/load-l1-from-excel.py              # dry-run (print summary)
  python3 scripts/load-l1-from-excel.py --execute     # execute against DB
"""
import sys, os, re, datetime, subprocess
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

EXCEL_PATH = os.path.expanduser(
    "~/Downloads/L1_Reference_Data_Export (UPDATED 313).xlsx"
)


def _git_root():
    """Find the git repo root (works in worktrees too)."""
    try:
        return Path(subprocess.check_output(
            ["git", "rev-parse", "--show-toplevel"], text=True, stderr=subprocess.DEVNULL
        ).strip())
    except (subprocess.CalledProcessError, FileNotFoundError):
        return None


def load_database_url():
    """Load DATABASE_URL from .env — checks script dir ancestors and git root."""
    d = Path(__file__).resolve().parent.parent
    candidates = [d, d.parent, d.parent.parent]
    git_root = _git_root()
    if git_root:
        candidates.append(git_root)
        # For worktrees, also check the main repo root
        main_root = git_root.parent.parent.parent
        if main_root != git_root and (main_root / ".env").exists():
            candidates.append(main_root)
    for candidate in candidates:
        env_file = candidate / ".env"
        if env_file.exists():
            for line in env_file.read_text().splitlines():
                if line.startswith("DATABASE_URL="):
                    return line.split("=", 1)[1].strip()
    return os.environ.get("DATABASE_URL")

# Tabs where header is in row 2 (row 1 is a comment)
HEADER_ROW_2 = {
    "origination_date_bucket_dim",
    "limit_rule",
    "limit_threshold",
    "country_dim",
    "metric_threshold",
}

# Excel column name → DB column name aliases
# The DB uses is_ prefix on boolean flag columns consistently
COLUMN_ALIASES = {
    "active_flag": "is_active_flag",
    "default_flag": "is_default_flag",
    "default_trigger_flag": "is_default_trigger_flag",
    "eligible_flag": "is_eligible_flag",
    "included_flag": "is_included_flag",
    "off_balance_sheet_flag": "is_off_balance_sheet_flag",
    "credit_event_trigger_flag": "is_credit_event_trigger_flag",
    "effective_from_date": "effective_start_date",
    "effective_to_date": "effective_end_date",
    "substatus_effective_to_date": "substatus_effective_end_date",
}

# All 49 tables to update (grouped by color for reporting)
YELLOW = [
    "dpd_bucket_dim", "internal_risk_rating_bucket_dim", "rating_mapping",
    "rating_grade_dim", "origination_date_bucket_dim", "limit_rule",
    "limit_threshold", "pricing_tier_dim", "region_dim", "country_dim",
    "utilization_status_dim", "collateral_eligibility_dim",
    "collateral_haircut_dim", "metric_threshold", "rating_scale_dim",
    "sccl_counterparty_group_member", "scenario_dim", "source_system_registry",
]
BLUE = [
    "collateral_portfolio", "context_dim", "counterparty_role_dim",
    "credit_status_dim", "instrument_identifier", "interest_rate_index_dim",
    "metric_definition_dim", "portfolio_dim", "rating_source",
    "risk_rating_tier_dim", "rule_registry", "sccl_counterparty_group",
    "validation_check_registry",
]
ORANGE = ["credit_event_type_dim", "crm_type_dim"]
GREEN = [
    "enterprise_business_taxonomy", "maturity_bucket_dim",
    "amendment_status_dim", "amendment_type_dim", "collateral_type",
    "crm_eligibility_dim", "date_dim", "date_time_dim",
    "default_definition_dim", "enterprise_product_taxonomy",
    "entity_type_dim", "exposure_type_dim", "fr2590_category_dim",
]
THEME = ["currency_dim", "industry_dim", "run_control"]

ALL_TABLES = YELLOW + BLUE + ORANGE + GREEN + THEME

COLOR_MAP = {}
for t in YELLOW: COLOR_MAP[t] = "YELLOW"
for t in BLUE: COLOR_MAP[t] = "BLUE"
for t in ORANGE: COLOR_MAP[t] = "ORANGE"
for t in GREEN: COLOR_MAP[t] = "GREEN"
for t in THEME: COLOR_MAP[t] = "THEME"


def read_excel_tab(wb, tab_name):
    """Read an Excel tab and return (columns, rows) handling quirks."""
    ws = wb[tab_name]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []

    # Determine header row
    if tab_name in HEADER_ROW_2:
        if len(all_rows) < 2:
            return [], []
        header_row = all_rows[1]
        data_rows = all_rows[2:]
    else:
        header_row = all_rows[0]
        data_rows = all_rows[1:]

    # region_dim: col 0 is "Remove" flag, actual headers start at col 1
    if tab_name == "region_dim":
        remove_col = [r[0] for r in data_rows]
        header_row = header_row[1:]
        data_rows_filtered = []
        for i, row in enumerate(data_rows):
            if remove_col[i] and str(remove_col[i]).strip().lower() == "remove":
                continue
            data_rows_filtered.append(row[1:])
        data_rows = data_rows_filtered

    # Clean headers: strip, lowercase, apply aliases, drop None/empty/Unnamed
    columns = []
    col_indices = []
    for i, h in enumerate(header_row):
        if h is None:
            continue
        h_clean = str(h).strip().lower()
        if not h_clean or h_clean.startswith("unnamed"):
            continue
        # Apply column name aliases (Excel name → DB name)
        h_clean = COLUMN_ALIASES.get(h_clean, h_clean)
        columns.append(h_clean)
        col_indices.append(i)

    # Extract data using only valid column indices
    cleaned_rows = []
    for row in data_rows:
        # Skip fully empty rows
        if all(v is None for v in row):
            continue
        vals = []
        for idx in col_indices:
            if idx < len(row):
                vals.append(row[idx])
            else:
                vals.append(None)
        cleaned_rows.append(vals)

    return columns, cleaned_rows


def get_db_columns(cur, table_name):
    """Get column info from information_schema for an L1 table."""
    cur.execute("""
        SELECT column_name, udt_name, is_nullable, column_default
        FROM information_schema.columns
        WHERE table_schema = 'l1' AND table_name = %s
        ORDER BY ordinal_position
    """, (table_name,))
    return {row[0]: {"udt_name": row[1], "nullable": row[2], "default": row[3]}
            for row in cur.fetchall()}


def get_pk_columns(cur, table_name):
    """Get primary key column(s) for an L1 table."""
    cur.execute("""
        SELECT kcu.column_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu
            ON tc.constraint_name = kcu.constraint_name
            AND tc.table_schema = kcu.table_schema
        WHERE tc.constraint_type = 'PRIMARY KEY'
        AND tc.table_schema = 'l1'
        AND tc.table_name = %s
        ORDER BY kcu.ordinal_position
    """, (table_name,))
    return [r[0] for r in cur.fetchall()]


def convert_value(val, udt_name, col_name):
    """Convert a Python/Excel value to a PostgreSQL-compatible value."""
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s.lower() == "nan" or s.lower() == "none":
        return None

    # Boolean
    if udt_name == "bool":
        if isinstance(val, bool):
            return val
        s_lower = s.lower()
        if s_lower in ("true", "1", "yes", "t"):
            return True
        if s_lower in ("false", "0", "no", "f"):
            return False
        return None

    # Integer types
    if udt_name in ("int4", "int2"):
        try:
            return int(float(s))
        except (ValueError, OverflowError):
            return None

    # Bigint
    if udt_name == "int8":
        try:
            return int(float(s))
        except (ValueError, OverflowError):
            return None

    # Numeric/decimal
    if udt_name == "numeric":
        # Handle percentage strings like "50%"
        if s.endswith("%"):
            try:
                return float(s[:-1]) / 100.0
            except ValueError:
                return None
        try:
            return float(s)
        except ValueError:
            return None

    # Date
    if udt_name == "date":
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        try:
            return datetime.datetime.strptime(s[:10], "%Y-%m-%d").date()
        except ValueError:
            return None

    # Timestamp
    if udt_name in ("timestamp", "timestamptz"):
        if isinstance(val, datetime.datetime):
            return val
        if isinstance(val, datetime.date):
            return datetime.datetime.combine(val, datetime.time.min)
        try:
            return datetime.datetime.fromisoformat(s)
        except ValueError:
            try:
                return datetime.datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None

    # Text/varchar - return as string
    if isinstance(val, datetime.datetime):
        return val.isoformat()
    if isinstance(val, datetime.date):
        return val.isoformat()
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return s


def main():
    execute = "--execute" in sys.argv
    db_url = load_database_url()
    if not db_url:
        print("ERROR: DATABASE_URL not found in .env or environment")
        sys.exit(1)

    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

    # Read all tab data
    tab_data = {}
    for table in ALL_TABLES:
        if table not in wb.sheetnames:
            print(f"  WARNING: Tab '{table}' not found in Excel, skipping")
            continue
        cols, rows = read_excel_tab(wb, table)
        if not cols:
            print(f"  WARNING: Tab '{table}' has no columns, skipping")
            continue
        tab_data[table] = (cols, rows)
        print(f"  Read {table}: {len(cols)} cols, {len(rows)} rows")
    wb.close()

    print(f"\nRead {len(tab_data)} tables from Excel")

    # Connect to DB
    conn = psycopg2.connect(db_url)
    try:
        cur = conn.cursor()

        # Get DB schema for each table and compute column intersection
        insert_plan = {}
        for table, (excel_cols, excel_rows) in tab_data.items():
            db_cols = get_db_columns(cur, table)
            if not db_cols:
                print(f"  WARNING: Table l1.{table} not found in DB, skipping")
                continue

            # Find intersection (columns in both Excel and DB)
            common_cols = [c for c in excel_cols if c in db_cols]
            excel_only = [c for c in excel_cols if c not in db_cols]

            if excel_only:
                print(f"  {table}: Excel-only cols (ignored): {excel_only}")

            # Build column index mapping
            col_idx_map = {c: excel_cols.index(c) for c in common_cols}
            col_types = {c: db_cols[c]["udt_name"] for c in common_cols}

            # Get PK columns for deduplication
            pk_cols = get_pk_columns(cur, table)

            # Convert all row values
            converted_rows = []
            seen_pks = set()
            dup_count = 0
            for row in excel_rows:
                converted = []
                for c in common_cols:
                    idx = col_idx_map[c]
                    raw = row[idx]
                    converted.append(convert_value(raw, col_types[c], c))
                # Skip rows where ALL values are None
                if all(v is None for v in converted):
                    continue
                # Deduplicate by PK
                if pk_cols:
                    pk_indices = [common_cols.index(c) for c in pk_cols if c in common_cols]
                    if pk_indices:
                        pk_val = tuple(converted[i] for i in pk_indices)
                        if pk_val in seen_pks:
                            dup_count += 1
                            continue
                        seen_pks.add(pk_val)
                converted_rows.append(tuple(converted))

            if dup_count:
                print(f"  {table}: WARNING: {dup_count} duplicate PK rows skipped")

            insert_plan[table] = {
                "columns": common_cols,
                "rows": converted_rows,
                "col_types": col_types,
            }

        if not execute:
            print("\n=== DRY RUN (add --execute to apply) ===")
            for table, plan in insert_plan.items():
                color = COLOR_MAP.get(table, "?")
                print(f"  [{color:6s}] l1.{table}: {len(plan['columns'])} cols, {len(plan['rows'])} rows")
            print(f"\nTotal: {len(insert_plan)} tables, {sum(len(p['rows']) for p in insert_plan.values())} rows")
            return

        # === EXECUTE ===
        print("\n=== EXECUTING against database ===")
        conn.rollback()
        conn.autocommit = False

        try:
            cur.execute("BEGIN;")

            # Collect ALL FK constraints referencing L1 tables (for drop/recreate)
            cur.execute("""
                SELECT
                    tc.constraint_name,
                    tc.table_schema AS child_schema,
                    tc.table_name AS child_table,
                    kcu.column_name AS child_column,
                    ccu.table_schema AS parent_schema,
                    ccu.table_name AS parent_table,
                    ccu.column_name AS parent_column
                FROM information_schema.table_constraints tc
                JOIN information_schema.key_column_usage kcu
                    ON tc.constraint_name = kcu.constraint_name
                    AND tc.table_schema = kcu.table_schema
                JOIN information_schema.constraint_column_usage ccu
                    ON tc.constraint_name = ccu.constraint_name
                    AND tc.table_schema = ccu.constraint_schema
                WHERE tc.constraint_type = 'FOREIGN KEY'
                AND ccu.table_schema = 'l1'
                ORDER BY tc.table_schema, tc.table_name
            """)
            fk_rows = cur.fetchall()

            # Build unique constraint definitions (handle multi-column FKs)
            fk_constraints = {}
            for cname, cs, ct, cc, ps, pt, pc in fk_rows:
                key = (cs, ct, cname)
                if key not in fk_constraints:
                    fk_constraints[key] = {
                        "child_schema": cs, "child_table": ct,
                        "constraint_name": cname,
                        "child_cols": [], "parent_schema": ps,
                        "parent_table": pt, "parent_cols": [],
                    }
                if cc not in fk_constraints[key]["child_cols"]:
                    fk_constraints[key]["child_cols"].append(cc)
                if pc not in fk_constraints[key]["parent_cols"]:
                    fk_constraints[key]["parent_cols"].append(pc)

            # Drop all FK constraints
            print(f"\n--- DROP {len(fk_constraints)} FK constraints ---")
            for key, fk in fk_constraints.items():
                cur.execute(
                    f'ALTER TABLE {fk["child_schema"]}.{fk["child_table"]} '
                    f'DROP CONSTRAINT IF EXISTS {fk["constraint_name"]};'
                )
            print(f"  Dropped {len(fk_constraints)} FK constraints")

            # Delete all rows
            print("\n--- DELETE phase ---")
            for table in insert_plan:
                cur.execute(f"DELETE FROM l1.{table};")
                print(f"  DELETED l1.{table}")

            # Insert phase
            print("\n--- INSERT phase ---")
            total_inserted = 0
            for table, plan in insert_plan.items():
                cols = plan["columns"]
                rows = plan["rows"]
                if not rows:
                    print(f"  l1.{table}: 0 rows (empty)")
                    continue

                col_list = ", ".join(f'"{c}"' for c in cols)
                insert_sql = f'INSERT INTO l1.{table} ({col_list}) VALUES %s'

                execute_values(cur, insert_sql, rows, page_size=500)

                color = COLOR_MAP.get(table, "?")
                print(f"  [{color:6s}] l1.{table}: {len(rows)} rows inserted")
                total_inserted += len(rows)

            # Commit the data changes first (before FK recreation)
            conn.commit()
            print("\nCOMMIT successful - data loaded!")

            # Now recreate FK constraints in a separate transaction
            print(f"\n--- RECREATE {len(fk_constraints)} FK constraints ---")
            fk_ok = 0
            fk_errors = []
            for key, fk in fk_constraints.items():
                child_cols = ", ".join(f'"{c}"' for c in fk["child_cols"])
                parent_cols = ", ".join(f'"{c}"' for c in fk["parent_cols"])
                sql = (
                    f'ALTER TABLE {fk["child_schema"]}.{fk["child_table"]} '
                    f'ADD CONSTRAINT {fk["constraint_name"]} '
                    f'FOREIGN KEY ({child_cols}) '
                    f'REFERENCES {fk["parent_schema"]}.{fk["parent_table"]} ({parent_cols});'
                )
                try:
                    cur.execute(sql)
                    conn.commit()
                    fk_ok += 1
                except Exception as e:
                    conn.rollback()
                    err_msg = str(e).strip().split("\n")[0]
                    fk_errors.append((fk["constraint_name"], err_msg))
            print(f"  Recreated {fk_ok}/{len(fk_constraints)} FK constraints")
            if fk_errors:
                print(f"  {len(fk_errors)} FK constraints FAILED (data integrity issues):")
                for name, err in fk_errors:
                    print(f"    {name}: {err}")
            print(f"Total: {total_inserted} rows inserted across {len(insert_plan)} tables")

        except Exception as e:
            conn.rollback()
            print(f"\nERROR: {e}")
            print("ROLLBACK - no changes applied")
            sys.exit(1)

        # === VERIFICATION ===
        print("\n=== VERIFICATION ===")
        mismatches = []
        for table, plan in insert_plan.items():
            expected = len(plan["rows"])
            cur.execute(f"SELECT count(*) FROM l1.{table}")
            actual = cur.fetchone()[0]
            status = "OK" if actual == expected else "MISMATCH"
            color = COLOR_MAP.get(table, "?")
            if status != "OK":
                mismatches.append((table, expected, actual))
                print(f"  [{color:6s}] l1.{table}: expected={expected}, actual={actual} *** {status} ***")
            else:
                print(f"  [{color:6s}] l1.{table}: {actual} rows OK")

        if mismatches:
            print(f"\n{len(mismatches)} MISMATCHES found!")
            for table, exp, act in mismatches:
                print(f"  l1.{table}: expected {exp}, got {act}")
        else:
            print(f"\nAll {len(insert_plan)} tables verified successfully!")

    finally:
        conn.close()


if __name__ == "__main__":
    main()
